#import pyaudio
import matplotlib.pyplot as plt
from scipy import signal
from scipy.io import wavfile
import os
import numpy as np
import random
from skimage.measure import block_reduce

#for excel editing
from openpyxl import Workbook


#To find the duration of wave file in seconds
import wave
import contextlib

#Keras imports
import keras
from keras.models import Sequential
from keras.layers import Dense, Conv2D, Dropout, Flatten, MaxPooling2D
from keras.models import model_from_json

import time
import datetime

os.environ['KMP_DUPLICATE_LIB_OK']  = 'True'

main_directory = 'C:\\Users\\rlewi\\Desktop\\audio learning\\train\\audio'
folder_names = [f for f in os.listdir(main_directory) if not 'background' in f]
folder_names.sort()
train_num = int(input("how many files would you like to train the model with? \n"))
test_num = int(input("how many files would you like to test the model with? \n"))
#train_list = np.empty((int(train_num), len(folder_names)), dtype=object)
wb = Workbook()
# creates active workbook var
ws = wb.active
# creates active worksheet var, set to zero by default and will grab first sheet


imheight = int(17)
imwidth = int(25)
k = int(0)

def save_model_to_disk(model):
    # serialize model to JSON
    model_json = model.to_json()
    with open("model.json", "w") as json_file:
        json_file.write(model_json)
    # serialize weights to HDF5
    model.save_weights("model.h5")
    print("Saved model to disk")


#Convert color image to grayscale
def rgb2gray(rgb):
    return np.dot(rgb[...,:3], [0.299, 0.587, 0.114])

#Normalize Gray colored image
def normalize_gray(array):
    return (array - array.min())/(array.max() - array.min())

# Function to find the duration of the wave file in seconds
def findDuration(fname):
    with contextlib.closing(wave.open(fname,'r')) as f:
        frames = f.getnframes()
        rate = f.getframerate()
        sw   = f.getsampwidth()
        chan = f.getnchannels()
        duration = frames / float(rate)
        #print("File:", fname, "--->",frames, rate, sw, chan)
        return duration

def graph_spectrogram(wav_file, nfft=512, noverlap=256):
    findDuration(wav_file)
    rate, data = wavfile.read(wav_file)
    #print("")
    fig,ax = plt.subplots(1)
    fig.subplots_adjust(left=0,right=1,bottom=0,top=1)
    ax.axis('off')
    pxx, freqs, bins, im = ax.specgram(x=data, Fs=rate, noverlap=noverlap, NFFT=nfft)
    ax.axis('off')
    plt.rcParams['figure.figsize'] = [0.75,0.5]
    #fig.savefig('sp_xyz.png', dpi=300, frameon='false')
    fig.canvas.draw()
    size_inches  = fig.get_size_inches()
    dpi          = fig.get_dpi()
    width, height = fig.get_size_inches() * fig.get_dpi()

    #print(size_inches, dpi, width, height)
    mplimage = np.frombuffer(fig.canvas.tostring_rgb(), dtype=np.uint8)
    #print("MPLImage Shape: ", np.shape(mplimage))
    imarray = np.reshape(mplimage, (int(height), int(width), 3))
    plt.close(fig)
    return imarray

print("train_num = " + str(train_num))
print("test_num = " + str(test_num))

for i in range(int(len(folder_names))):
	# goes to the column for the respective audio folder
	current_folder = main_directory + '\\' + folder_names[i]
	# sets the directory
	file_list = [f for f in os.listdir(current_folder) if '.wav' in f]
	# creates a list of the wav files in the folder
	ws.cell(row=1, column=i+1).value = folder_names[i]
	# places the name of the folder at the top of each column
	for j in range(int(train_num+test_num)):
		# iterates through the rows
		#train_list[j,i] = random.choice(file_list)
		ws.cell(row=j+2, column=i+1).value = random.choice(file_list)
		# populate the rows with the respective file names

spreadsheetname = 'train_list.xlsx'
if os.path.exists(main_directory + "\\" + spreadsheetname):
	spreadsheetname = input("what would you like to name the spreadsheet? ") + ".xlsx"
wb.save(spreadsheetname)

y_train = np.zeros(len(folder_names)*int(train_num))
# size of the list used to train the neural net
x_train = np.zeros((int(len(y_train)), imheight, imwidth))
# shape of the training data set
y_test = np.zeros(len(folder_names)*int(test_num))
# size of the list used to test the neural net
x_test = np.zeros((int(len(y_test)), imheight, imwidth))
# shape of the testing data set


#modify later to makes sure there's no duplicate entries being used,
#designate the last X percent of rows for use as testing data
#maybe figure out how to make the top row bold and other "human readable touches"


#print("Size of Training Data: ", np.shape(x_train))
#print("Size of Training Labels: ", np.shape(y_train))
#print("Size of Test Data: ", np.shape(x_test))
#print("Size of Test Labels: ", np.shape(y_test))


#maybe put an option to start here after excel file has been generated


#create a list that contains a number or row/column representing each piece of data
#use the random selection function to pull a value from it
#remove that value from the list (see if it's possible to do this from the middle)
#rinse and repeat

print("length of folder names: " + str(len(folder_names)))
print("number of files to be processed per folder: " + str(train_num+test_num))
for i in range(int(len(folder_names))):
	# goes to the column for the respective audio folder
	folder_name = ws.cell(row=1, column=i+1).value
	for j in range(train_num+test_num):
		# iterates through the rows
		file_name = ws.cell(row=j+2, column=i+1).value
		file_path = str(main_directory) + "\\" + str(folder_name) + "\\" +str(file_name) 
		print(str(j))
		spectrogram = graph_spectrogram(file_path)
		graygram = rgb2gray(spectrogram)
		normgram = normalize_gray(graygram)
		norm_shape = normgram.shape
		if(norm_shape[0]>150):
			continue
		redgram = block_reduce(normgram, block_size = (3,3), func = np.mean)
		if j < int(train_num):
			# check to make sure the split is right
			x_train[j,:,:] = redgram
			y_train[j] = k
			print(folder_name + ": Training Data Progress = {:2.1%}".format(float(j+1) / int(train_num), end="\r"))
		else:
			x_test[j,:,:] = redgram
			y_test[j] = k
			print(folder_name + ": Testing Data Progress = {:2.1%}".format(float(j+1-int(train_num)) / int(test_num), end="\r"))
		if j == train_num:
			k = 0
		k = k + 1
	




#create keras model
num_classes = 30

x_train = x_train.reshape(x_train.shape[0], imheight, imwidth, 1)
y_train = keras.utils.to_categorical(y_train, num_classes)

x_test = x_train.reshape(x_test.shape[0], imheight, imwidth, 1)
y_test = keras.utils.to_categorical(y_test, num_classes)

print("x and y training/testing data done and formatted, starting keras sequential model function now")

input_shape = (imheight, imwidth, 1)
batch_size = 4
epochs = 1

model = Sequential()
model.add(Conv2D(32, kernel_size=(3, 3), activation='relu', input_shape=input_shape))
model.add(Conv2D(64, kernel_size=(3, 3), activation='relu'))
model.add(MaxPooling2D(pool_size=(2, 2)))
model.add(Dropout(0.25))
model.add(Flatten())
model.add(Dense(128, activation='relu'))
model.add(Dropout(0.5))
model.add(Dense(num_classes, activation='softmax'))

model.compile(loss=keras.losses.categorical_crossentropy, optimizer=keras.optimizers.adam(), metrics=['accuracy'])
print(model.summary())

model.fit(x_train, y_train, batch_size=4, epochs=10, verbose=1, validation_data=(x_test, y_test))

save_model_to_disk(model)